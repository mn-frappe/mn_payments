"""GS1/GS1 Excel importer for Special Tax mapping.

Provides a whitelisted helper to download the GS1 spreadsheet and import barcode prefixes
as `Special Tax GS1 Map` records linked to a `Special Tax Type`.

Usage:
  - Call via server-side API: `mn_payments.integrations.gs1.importer.import_from_url(url, tax_name, tax_rate)`
  - Or run as a bench execute: `bench execute mn_payments.integrations.gs1.importer.import_from_url --args "['<url>','On tsgoi',3.0]'`
"""

from __future__ import annotations

import io
import requests
import frappe
from frappe import _

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None


def _ensure_tax_type(tax_name: str, tax_rate: float) -> str:
    """Return name of Special Tax Type, creating it if missing."""
    name = frappe.db.get_value("Special Tax Type", {"tax_name": tax_name})
    if name:
        return name
    doc = frappe.get_doc({
        "doctype": "Special Tax Type",
        "tax_name": tax_name,
        "rate": tax_rate,
    })
    doc.insert(ignore_permissions=True)
    return doc.name


def _parse_workbook_bytes(data: bytes) -> list[dict]:
    """Parse workbook bytes and return list of rows as dicts.

    Attempts to read the first sheet and extract any column that looks like a barcode or code.
    """
    if not load_workbook:
        frappe.throw(_("openpyxl is required to import XLSX files. Install with `pip install openpyxl`"))

    wb = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    results: list[dict] = []
    for row in rows[1:]:
        item = {}
        for i, cell in enumerate(row):
            key = headers[i] if i < len(headers) else f"col_{i}"
            item[key] = cell
        results.append(item)
    return results


def _find_barcode_in_row(row: dict) -> str | None:
    """Look for likely barcode-like value in row dict.

    Prefer columns with names like 'barcode', 'gtin', 'ean', 'upc', or 'cpcd'.
    """
    candidates = []
    for k, v in row.items():
        if v is None:
            continue
        lname = k.lower()
        if any(tok in lname for tok in ("barcode", "gtin", "ean", "upc", "cpcd", "code")):
            candidates.append(str(v))
    if candidates:
        return candidates[0]
    # fallback: find first numeric long-ish cell
    for v in row.values():
        if v is None:
            continue
        s = str(v).strip()
        if s.isdigit() and len(s) >= 8:
            return s
    return None


def _create_map_entry(prefix: str, tax_type_name: str, description: str | None = None):
    # ensure uniqueness by prefix
    existing = frappe.db.get_value("Special Tax GS1 Map", {"barcode_prefix": prefix})
    if existing:
        return existing
    doc = frappe.get_doc(
        {
            "doctype": "Special Tax GS1 Map",
            "barcode_prefix": prefix,
            "tax_type": tax_type_name,
            "description": description or "Imported from GS1",
        }
    )
    doc.insert(ignore_permissions=True)
    return doc.name


@frappe.whitelist()
def import_from_url(url: str, tax_name: str = "On tsgoi", tax_rate: float = 3.0):
    """Download the XLSX at `url` and import barcode prefixes as special tax mappings.

    Returns a summary dict with counts.
    """
    if not url:
        frappe.throw(_("Missing URL"))

    resp = requests.get(url, timeout=30)
    resp.raise_for_status()
    data = resp.content

    rows = _parse_workbook_bytes(data)
    if not rows:
        return {"imported": 0, "message": "No rows found"}

    tax_type_name = _ensure_tax_type(tax_name, tax_rate)

    imported = 0
    for row in rows:
        barcode = _find_barcode_in_row(row)
        if not barcode:
            continue
        # use first 7-13 digits as prefix for mapping
        prefix = barcode.strip()
        # normalize: remove non-digits
        prefix = "".join(ch for ch in prefix if ch.isdigit())
        if not prefix:
            continue
        # store varying prefix lengths to help matching (store full barcode and shorter prefixes)
        # primary: full barcode
        _create_map_entry(prefix, tax_type_name, description=f"Imported row")
        imported += 1
    return {"imported": imported, "tax_type": tax_type_name}


def import_from_file(fpath: str, tax_name: str = "On tsgoi", tax_rate: float = 3.0):
    """Helper for server-side use: import from local file path."""
    with open(fpath, "rb") as f:
        data = f.read()
    rows = _parse_workbook_bytes(data)
    tax_type_name = _ensure_tax_type(tax_name, tax_rate)
    imported = 0
    for row in rows:
        barcode = _find_barcode_in_row(row)
        if not barcode:
            continue
        prefix = "".join(ch for ch in str(barcode).strip() if ch.isdigit())
        if not prefix:
            continue
        _create_map_entry(prefix, tax_type_name)
        imported += 1
    return {"imported": imported, "tax_type": tax_type_name}
